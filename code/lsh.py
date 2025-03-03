from ont_fast5_api.fast5_interface import get_fast5_file
from tqdm import tqdm
import matplotlib.pyplot as plt
import pickle
import torch
from uncalled4 import PoreModel, Config, EventDetector, SignalProcessor
import numpy as np
import pandas as pd
import sequence_to_signal
import cam_search
from cam_array import CAMArray
from event_processor import EventProcessor, fast5_id_list
from cam_process import process_sample, process_sample_contamination, process_sample_variation, process_sample_contamination_variation
from cam_process import update_position, update_position_contamination, update_position_variation, update_position_contamination_variation
from cam_process import process_location
from openpyxl import load_workbook
import argparse
from pafstats import run, calculate

import argparse

def main():
    parser = argparse.ArgumentParser(description="A program that uses Threshold and Sample_number from command line.")

    parser.add_argument('--read_number', type=int, required=True, help='Set the read number')
    parser.add_argument('--threshold', type=float, required=True, help='Set the threshold value')
    parser.add_argument('--sample_number', type=int, required=True, help='Set the sample number')
    parser.add_argument('--std', type=float, required=True, help='Set the variation of memristor')
    parser.add_argument('--file_path', type=str, required=True, help='Path to the fast5 file')

    args = parser.parse_args()

    _Threshold = args.threshold
    Sample_number = args.sample_number
    std_rram = args.std
    _read_number = args.read_number
    Fast5_file = args.file_path

    return _read_number, _Threshold, Sample_number, std_rram, Fast5_file

if __name__ == '__main__':
    read_number, _Threshold, sample_number, std, fast5_file = main()
    std = std*1e-6
    print("reads:", read_number)
    print("Threshold:", _Threshold)
    print("Sample_number:",sample_number)
    print("std:",std)
    print("fast5_file:", fast5_file)


# Read the main sequence
main_sequence_path = '../dataset/sarscov2.fna'
main_sequences = sequence_to_signal.read_fasta(main_sequence_path)
main_sequence_length = sequence_to_signal.get_sequence_length(main_sequence_path)

# Read the complement sequence
complement_sequence_path = '../dataset/sarscov2_complement_true.fna'
complement_sequences = sequence_to_signal.read_fasta(complement_sequence_path)
complement_sequence_length = sequence_to_signal.get_sequence_length(complement_sequence_path)

# print(f"Length of main sequence: {main_sequence_length}")
# print(f"Length of complement sequence: {complement_sequence_length}")

# Define the model file path
model_file_path = "../dataset/kmer_model/template_median68pA.model" #r94
df = sequence_to_signal.load_data(model_file_path)

# Retrieve k-mer means and standard deviations for the original reference sequence
kmer_means_raw = sequence_to_signal.get_kmer_properties(main_sequences, df)
kmer_means = sequence_to_signal.replace_continuous_signal(kmer_means_raw)

# Retrieve k-mer means and standard deviations for the complement reference sequence
kmer_means_comp_raw = sequence_to_signal.get_kmer_properties(complement_sequences, df)
kmer_means_comp = sequence_to_signal.replace_continuous_signal(kmer_means_comp_raw)


# print(f"K-mer means for the original sequence: {kmer_means}, Standard deviation: {kmer_std}")
# print(f"K-mer means for the complement sequence: {kmer_means_comp}, Standard deviation: {kmer_std_comp}")
zero_count = sequence_to_signal.count_zeros(kmer_means_comp)

loaded_conductance = np.load('random_conductance/raw_conductance_simulation.npy')

flat_list = loaded_conductance.flatten().tolist()
template= flat_list[0:2560]
difference_list_raw = [template[i] - template[i + 1] for i in range(0, len(template) - 1, 2)]

random_conductance = np.array(difference_list_raw).reshape(10, 128)
random_matrix_tensor = torch.tensor(random_conductance, device='cuda').float()

col = 10
LSH_col = 128
k = 10
sub_array_row = 400

cam_processor = CAMArray(col=10, LSH_col=128, k=10, sub_array_row=400,
                             kmer_level_means_shift=kmer_means,
                             kmer_level_means_shift_comp=kmer_means_comp, random_matrix=random_conductance, device='cuda')
cam_processor.initialize_matrices()
ref_array_tensor, ref_array_comp_tensor = cam_processor.process_LSH()
cam_processor.print_shapes()

# Calculate number of blocks(sub_array)
n_blocks = cam_processor.calculate_sub_arrays(ref_array_tensor)

# Two parameter for event
sp, sp1 = EventProcessor()
#read read_id from list samples here
read_id = fast5_id_list(fast5_file)

position = []
direction = []
search_time = []
vote_location = []

torch.manual_seed(42)

gon = torch.normal(mean=146.79*1e-6, std=std, size=ref_array_tensor.shape, device = 'cuda')  # experimental
goff = abs(torch.normal(mean=4.33*1e-6, std=std, size=ref_array_tensor.shape,  device = 'cuda'))  # experimental


# three parameter for process
difference = [3, 3.5, 4]
# Threshold = _Threshold *146.79 + (LSH_col-_Threshold)*4.33
Threshold = _Threshold *146.79*1e-6 + (LSH_col - _Threshold)*4.33*1e-6


# read_number = 10000

for i in tqdm(range(0,read_number), mininterval=100):
    _read_id = read_id[i]
    final_location, dir, _search_time, _votes = process_sample_variation(gon, goff, sample_number, sp, _read_id, fast5_file, random_matrix_tensor, ref_array_tensor, ref_array_comp_tensor, col, Threshold, sub_array_row, n_blocks, difference, device = 'cuda')
    
    position.append(final_location)
    direction.append(dir)
    search_time.append(_search_time)
    vote_location.append(_votes)

N_count = sum(1 for item in position if item == 'N')

index = [i for i, x in enumerate(position) if x == 'N']
update_position_variation(gon, goff, sample_number, position, direction, search_time, vote_location, sp1, read_id, fast5_file, index, [4, 3], random_matrix_tensor, ref_array_tensor, ref_array_comp_tensor, col, Threshold, sub_array_row, n_blocks, device = 'cuda')

N_count = sum(1 for item in position if item == 'N')

# Create a DataFrame from the lists
df = pd.DataFrame({
    'Column1': search_time,
    'Column2': vote_location,
})

file_path_csv = "result/{}_cov_votes_threshold_{}_std_{}.csv".format(sample_number, _Threshold, std)

# Save the DataFrame to a CSV file
# file_path = 'result/cov_mapping_threshold_1.csv'
df.to_csv(file_path_csv, index=False)


_position=position
_direction=direction
_direction = ['*' if x == 'N' else x for x in _direction]

low_boundary=[]
high_boundary=[]
_position=np.array(_position)
low_boundary,high_boundary= process_location(sample_number, sp, low_boundary, high_boundary, read_id, read_number, _position, fast5_file, sub_array_row)

data_list = _direction

# load template
# template_file_path = 'result_template/test_10k_template.xlsx'
if fast5_file == '../dataset/SP1-mapped500.fast5':
    template_file_path = 'result_template/test_SP1_500_template.xlsx'
else:
    template_file_path = 'result_template/test_10k_template.xlsx'

wb = load_workbook(template_file_path)
ws = wb.active  

for index, item in enumerate(data_list, start=1):  
    ws[f'E{index}'].value = item

for index, item in enumerate(low_boundary, start=1):  
    ws[f'H{index}'].value = item

for index, item in enumerate(high_boundary, start=1):  
    ws[f'I{index}'].value = item

# save
output_file_path = "result/{}_test_LSH_rm_threshold_{}_std_{}.xlsx".format(sample_number, _Threshold, std)

wb.save(output_file_path)

output_file_path_txt = "result/{}_test_LSH_rm_threshold_{}_std_{}".format(sample_number, _Threshold, std)

# write
with open(output_file_path_txt, 'w', encoding='utf-8') as f:
    for row in ws.iter_rows(values_only=True):
        row_data = '\t'.join(map(str, row))
        f.write(row_data + '\n')

args = argparse.Namespace(
    infile=output_file_path_txt,  
    max_reads=read_number,  
    ref_paf="result_template/minimap2_sars2.paf",  
    annotate=False  
)

run(args)